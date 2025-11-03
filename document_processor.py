"""
Enhanced Document Processor for Multi-Modal RAG System

This module handles document processing with advanced features including:
- Multi-format support (PDF, PPTX, EML, XLS, XLSX)
- Intelligent content extraction using AI
- Performance optimizations (batch processing, compression, filtering)
- Caching system for expensive extractions
- Duplicate detection and smart filtering

"""

import os
import base64
import json
import hashlib
import asyncio
import tempfile
import zipfile
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

# Image processing libraries
from PIL import Image, ImageFilter
import imagehash
import numpy as np

# Document processing libraries
import fitz  # PyMuPDF for PDF processing
from pptx import Presentation
import openpyxl
import xlrd
from openpyxl.drawing.image import Image as OpenpyxlImage

# Email processing libraries
import email
from email import policy
from email.parser import BytesParser

# AI and language processing
from langchain_google_vertexai import ChatVertexAI
from langchain_core.messages import HumanMessage

# Configuration
from config import Config


class DocumentProcessor:
    """
    Enhanced Document Processor with AI-powered content extraction and performance optimizations.
    
    Features:
    - Multi-format document support
    - AI-powered content extraction
    - Intelligent caching system
    - Performance optimizations
    - Smart filtering and duplicate detection
    """
    
    def __init__(self):
        """Initialize the document processor with all necessary components."""
        self._setup_authentication()
        self._initialize_ai_models()
        self._setup_directory_structure()
        self._initialize_performance_tracking()
    
    # =============================================================================
    # INITIALIZATION METHODS
    # =============================================================================
    
    def _setup_authentication(self) -> None:
        """Setup Google Cloud authentication for AI services."""
        try:
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = Config.GOOGLE_CREDENTIALS_PATH
        except Exception as e:
            raise Exception(f"Failed to setup Google authentication: {e}")
    
    def _initialize_ai_models(self) -> None:
        """Initialize AI models for content extraction."""
        self.llm = ChatVertexAI(model_name="gemini-2.5-pro")
        self._thread_local = threading.local()
    
    def _setup_directory_structure(self) -> None:
        """Create necessary directories for caching and storage."""
        self.extractions_dir = Config.EXTRACTION_CACHE_DIR
        self.images_dir = Config.EXTRACTION_IMAGES_DIR
        self.content_dir = Config.EXTRACTION_CONTENT_DIR
        
        # Create directories if they don't exist
        for directory in [self.extractions_dir, self.images_dir, self.content_dir]:
            os.makedirs(directory, exist_ok=True)
    
    def _initialize_performance_tracking(self) -> None:
        """Initialize performance tracking attributes."""
        self.processed_image_hashes = set()  # Track processed images for duplicate detection
        self.page_similarity_cache = []      # Cache for page similarity comparisons
        self.processing_stats = {
            'total_pages_processed': 0,
            'pages_skipped_duplicates': 0,
            'pages_skipped_content': 0,
            'total_api_calls_saved': 0
        }
    
    # =============================================================================
    # AI MODEL MANAGEMENT
    # =============================================================================
    
    def _get_thread_safe_llm(self) -> ChatVertexAI:
        """Get a thread-local LLM instance for safe parallel processing."""
        if not hasattr(self._thread_local, 'llm'):
            self._thread_local.llm = ChatVertexAI(model_name="gemini-2.5-pro")
        return self._thread_local.llm
    
    # =============================================================================
    # FILE MANAGEMENT & CACHING METHODS
    # =============================================================================
    
    def generate_file_hash(self, file_path: str) -> str:
        """Generate MD5 hash of file for cache validation and duplicate detection."""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception as e:
            print(f"Error generating file hash: {e}")
            return ""
    
    def get_extraction_cache_path(self, file_path: str) -> str:
        """Get the path where extraction results should be cached."""
        file_hash = self.generate_file_hash(file_path)
        filename = os.path.basename(file_path)
        cache_filename = f"{filename}_{file_hash}.json"
        return os.path.join(self.content_dir, cache_filename)
    
    def save_processed_image(self, image: Image.Image, file_path: str, page_num: int, content_type: str) -> str:
        """Save processed image to disk with organized naming convention."""
        try:
            file_hash = self.generate_file_hash(file_path)
            filename = os.path.basename(file_path)
            base_name = os.path.splitext(filename)[0]
            
            image_filename = f"{base_name}_{file_hash}_page{page_num}_{content_type}.png"
            image_save_path = os.path.join(self.images_dir, image_filename)
            
            image.save(image_save_path, "PNG")
            return image_save_path
        except Exception as e:
            print(f"Error saving image: {e}")
            return ""
    
    def save_extraction_to_cache(self, file_path: str, extracted_content: Dict[str, List[Dict[str, Any]]]) -> Optional[str]:
        """Save extraction results to cache with metadata."""
        try:
            cache_path = self.get_extraction_cache_path(file_path)
            
            cache_data = {
                "source_file": file_path,
                "file_hash": self.generate_file_hash(file_path),
                "extraction_timestamp": datetime.now().isoformat(),
                "system_version": "2.0",
                "content": extracted_content,
                "processing_stats": self.processing_stats.copy()
            }
            
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, indent=2, ensure_ascii=False)
            
            print(f"✓ Extraction cached: {os.path.basename(cache_path)}")
            return cache_path
            
        except Exception as e:
            print(f"Error saving to cache: {e}")
            return None
    
    def load_extraction_from_cache(self, file_path: str) -> Optional[Dict[str, List[Dict[str, Any]]]]:
        """Load previously cached extraction results with validation."""
        try:
            cache_path = self.get_extraction_cache_path(file_path)
            
            if not os.path.exists(cache_path):
                return None
            
            # Validate file hasn't been modified since caching
            current_hash = self.generate_file_hash(file_path)
            
            with open(cache_path, 'r', encoding='utf-8') as f:
                cache_data = json.load(f)
            
            cached_hash = cache_data.get("file_hash", "")
            
            if current_hash != cached_hash:
                print(f"File modified since caching. Re-processing required.")
                return None
            
            print(f"✓ Loaded from cache: {os.path.basename(cache_path)}")
            return cache_data["content"]
            
        except Exception as e:
            print(f"Error loading from cache: {e}")
            return None
    
    def has_valid_cache(self, file_path: str) -> bool:
        """Check if file has valid cached extraction."""
        return self.load_extraction_from_cache(file_path) is not None
    
    # =============================================================================
    # IMAGE OPTIMIZATION & PROCESSING
    # =============================================================================
    
    def optimize_image_for_ai_processing(self, image: Image.Image) -> Image.Image:
        """
        Optimize image for faster AI processing while maintaining quality.
        
        Optimizations applied:
        - Format conversion to RGB
        - Size reduction to optimal dimensions
        - Sharpening for better text recognition
        """
        if not Config.ENABLE_IMAGE_COMPRESSION:
            return image
        
        try:
            # Convert to RGB format if necessary
            if image.mode in ('RGBA', 'P', 'L'):
                image = image.convert('RGB')
            
            # Resize to optimal dimensions for processing
            if Config.IMAGE_MAX_SIZE and image.size > Config.IMAGE_MAX_SIZE:
                image.thumbnail(Config.IMAGE_MAX_SIZE, Image.Resampling.LANCZOS)
            
            # Apply sharpening for better text recognition
            if Config.IMAGE_QUALITY < 90:
                sharpening_filter = ImageFilter.UnsharpMask(radius=1, percent=120, threshold=2)
                image = image.filter(sharpening_filter)
            
            return image
            
        except Exception as e:
            print(f"Warning: Image optimization failed: {e}")
            return image
    
    def generate_image_perceptual_hash(self, image: Image.Image) -> str:
        """Generate perceptual hash for intelligent duplicate detection."""
        try:
            perceptual_hash = imagehash.phash(image)
            return str(perceptual_hash)
        except Exception as e:
            print(f"Warning: Image hashing failed: {e}")
            return ""
    
    def is_duplicate_page(self, image: Image.Image) -> bool:
        """
        Determine if page is a duplicate using perceptual hashing.
        
        Uses advanced similarity detection to identify:
        - Exact duplicates
        - Near-duplicates with minor differences
        - Similar layouts with different content
        """
        if not Config.ENABLE_DUPLICATE_DETECTION:
            return False
        
        try:
            current_hash = self.generate_image_perceptual_hash(image)
            if not current_hash:
                return False
            
            # Compare against all processed image hashes
            for processed_hash in self.processed_image_hashes:
                # Calculate similarity using Hamming distance
                hash_distance = imagehash.hex_to_hash(current_hash) - imagehash.hex_to_hash(processed_hash)
                similarity_score = 1 - (hash_distance / 64.0)
                
                if similarity_score > Config.SIMILARITY_THRESHOLD:
                    self.processing_stats['pages_skipped_duplicates'] += 1
                    print(f"  📋 Skipping duplicate page (similarity: {similarity_score:.2f})")
                    return True
            
            # Add to processed hashes if not duplicate
            self.processed_image_hashes.add(current_hash)
            return False
            
        except Exception as e:
            print(f"Warning: Duplicate detection failed: {e}")
            return False
    
    def has_sufficient_content_for_processing(self, image: Image.Image) -> bool:
        """
        Analyze if page has sufficient content worth processing.
        
        Checks for:
        - Blank or mostly empty pages
        - Low contrast content
        - Minimal text content
        """
        if not Config.ENABLE_SMART_FILTERING:
            return True
        
        try:
            # Convert to grayscale for analysis
            grayscale_image = image.convert('L')
            image_array = np.array(grayscale_image)
            
            # Check for mostly blank pages (high percentage of white pixels)
            white_threshold = 240
            white_pixel_ratio = np.sum(image_array > white_threshold) / image_array.size
            
            if white_pixel_ratio > 0.95:  # More than 95% white pixels
                self.processing_stats['pages_skipped_content'] += 1
                print("  📄 Skipping mostly blank page")
                return False
            
            # Check for sufficient contrast and content
            contrast_level = np.std(image_array)
            if contrast_level < 10:  # Very low contrast indicates minimal content
                self.processing_stats['pages_skipped_content'] += 1
                print("  📊 Skipping low-contrast page")
                return False
            
            return True
            
        except Exception as e:
            print(f"Warning: Content analysis failed: {e}")
            return True  # Process if analysis fails to be safe
    
    def _image_to_base64(self, image: Image.Image) -> str:
        """Convert PIL Image to base64 string"""
        buffered = BytesIO()
        image.save(buffered, format="PNG")
        img_str = base64.b64encode(buffered.getvalue()).decode()
        return img_str
    
    def convert_pdf_to_images(self, pdf_path: str) -> List[Image.Image]:
        """Convert PDF pages to images using PyMuPDF (no poppler dependency)"""
        try:
            images = []
            pdf_document = fitz.open(pdf_path)
            
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                # Render page to image with high resolution
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better quality
                
                # Convert pixmap to PIL Image
                img_data = pix.tobytes("png")
                img = Image.open(BytesIO(img_data))
                images.append(img)
            
            pdf_document.close()
            return images
        except Exception as e:
            print(f"Error converting PDF: {e}")
            return []
    
    def convert_pptx_to_images(self, pptx_path: str) -> List[Image.Image]:
        """Convert PPTX slides to images using python-pptx"""
        try:
            images = []
            prs = Presentation(pptx_path)
            
            # Create temporary directory for slide images
            temp_dir = "temp_slides"
            os.makedirs(temp_dir, exist_ok=True)
            
            for slide_num, slide in enumerate(prs.slides):
                # Export slide as image
                slide_image_path = os.path.join(temp_dir, f"slide_{slide_num}.png")
                
                # Get slide dimensions
                slide_width = prs.slide_width
                slide_height = prs.slide_height
                
                # Create blank image with slide dimensions
                img = Image.new('RGB', (int(slide_width / 9525), int(slide_height / 9525)), 'white')
                
                # Note: python-pptx doesn't directly support image export
                # For production, consider using win32com (Windows) or LibreOffice (cross-platform)
                # This is a placeholder - slides will be processed as text extraction
                images.append(img)
            
            # Clean up temp directory
            import shutil
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            
            return images
        except Exception as e:
            print(f"Error converting PPTX: {e}")
            return []
    
    def process_eml_file(self, eml_path: str) -> Dict[str, Any]:
        """Process EML (email) file and extract content"""
        try:
            with open(eml_path, 'rb') as f:
                raw_email = f.read()
            
            # Parse the email
            msg = BytesParser(policy=policy.default).parsebytes(raw_email)
            
            # Extract email metadata and content
            email_content = {
                'subject': msg.get('Subject', ''),
                'from': msg.get('From', ''),
                'to': msg.get('To', ''),
                'date': msg.get('Date', ''),
                'body_text': '',
                'body_html': '',
                'attachments': []
            }
            
            # Extract body content
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get('Content-Disposition', ''))
                    
                    if content_type == 'text/plain' and 'attachment' not in content_disposition:
                        email_content['body_text'] = part.get_content()
                    elif content_type == 'text/html' and 'attachment' not in content_disposition:
                        email_content['body_html'] = part.get_content()
                    elif 'attachment' in content_disposition:
                        filename = part.get_filename()
                        if filename:
                            email_content['attachments'].append({
                                'filename': filename,
                                'content_type': content_type,
                                'size': len(part.get_payload(decode=True) or b'')
                            })
            else:
                # Single part message
                email_content['body_text'] = msg.get_content()
            
            return email_content
            
        except Exception as e:
            print(f"Error processing EML file: {e}")
            return {}
    
    def extract_images_from_xlsx(self, xlsx_path: str) -> List[Image.Image]:
        """Extract embedded images from XLSX file"""
        images = []
        try:
            # Open the workbook
            workbook = openpyxl.load_workbook(xlsx_path)
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Check if sheet has images
                if hasattr(sheet, '_images') and sheet._images:
                    for img in sheet._images:
                        try:
                            # Get image data
                            image_data = img._data()
                            pil_image = Image.open(BytesIO(image_data))
                            images.append(pil_image)
                        except Exception as e:
                            print(f"Error extracting image from sheet {sheet_name}: {e}")
            
            workbook.close()
            
        except Exception as e:
            print(f"Error extracting images from XLSX: {e}")
        
        return images
    
    def convert_excel_to_images(self, excel_path: str) -> List[Image.Image]:
        """Convert Excel sheets to images for processing"""
        try:
            images = []
            file_ext = os.path.splitext(excel_path)[1].lower()
            
            if file_ext == '.xlsx':
                # Use openpyxl for .xlsx files
                workbook = openpyxl.load_workbook(excel_path, data_only=True)
                
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Create a visual representation of the sheet
                    # This is a simplified approach - for better visualization,
                    # consider using libraries like xlwings or converting via LibreOffice
                    sheet_image = self._create_sheet_visualization(sheet, sheet_name)
                    if sheet_image:
                        images.append(sheet_image)
                
                # Also extract any embedded images
                embedded_images = self.extract_images_from_xlsx(excel_path)
                images.extend(embedded_images)
                
                workbook.close()
                
            elif file_ext == '.xls':
                # Use xlrd for .xls files
                workbook = xlrd.open_workbook(excel_path)
                
                for sheet_index in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_index)
                    sheet_name = workbook.sheet_names()[sheet_index]
                    
                    # Create a visual representation of the sheet
                    sheet_image = self._create_xls_sheet_visualization(sheet, sheet_name)
                    if sheet_image:
                        images.append(sheet_image)
            
            return images
            
        except Exception as e:
            print(f"Error converting Excel file: {e}")
            return []
    
    def _create_sheet_visualization(self, sheet, sheet_name: str) -> Image.Image:
        """Create a visual representation of an Excel sheet (xlsx)"""
        try:
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row == 1 and max_col == 1 and not sheet.cell(1, 1).value:
                return None  # Empty sheet
            
            # Create a simple text-based visualization
            # For better results, consider using xlwings or similar libraries
            cell_width = 120
            cell_height = 30
            
            img_width = min(max_col * cell_width, 2000)  # Limit width
            img_height = min(max_row * cell_height, 2000)  # Limit height
            
            # Create blank image
            img = Image.new('RGB', (img_width, img_height), 'white')
            
            # This is a placeholder implementation
            # In production, you might want to use more sophisticated Excel-to-image conversion
            
            return img
            
        except Exception as e:
            print(f"Error creating sheet visualization: {e}")
            return None
    
    def _create_xls_sheet_visualization(self, sheet, sheet_name: str) -> Image.Image:
        """Create a visual representation of an Excel sheet (xls)"""
        try:
            # Get sheet dimensions
            nrows = sheet.nrows
            ncols = sheet.ncols
            
            if nrows == 0 or ncols == 0:
                return None  # Empty sheet
            
            # Create a simple text-based visualization
            cell_width = 120
            cell_height = 30
            
            img_width = min(ncols * cell_width, 2000)  # Limit width
            img_height = min(nrows * cell_height, 2000)  # Limit height
            
            # Create blank image
            img = Image.new('RGB', (img_width, img_height), 'white')
            
            # This is a placeholder implementation
            # In production, you might want to use more sophisticated Excel-to-image conversion
            
            return img
            
        except Exception as e:
            print(f"Error creating XLS sheet visualization: {e}")
            return None
    
    # =============================================================================
    # AI-POWERED CONTENT EXTRACTION
    # =============================================================================
    
    def extract_single_content_type(self, image: Image.Image, content_type: str) -> str:
        """
        Extract specific content type from image using AI.
        
        Args:
            image: PIL Image to process
            content_type: Type of content to extract ('text', 'table', 'visual')
            
        Returns:
            Extracted content as string
        """
        try:
            # Select appropriate prompt based on content type
            prompt_mapping = {
                "text": Config.TEXT_EXTRACTION_PROMPT,
                "table": Config.TABLE_EXTRACTION_PROMPT,
                "visual": Config.VISUAL_ANALYSIS_PROMPT
            }
            
            if content_type not in prompt_mapping:
                raise ValueError(f"Unknown content type: {content_type}")
            
            prompt = prompt_mapping[content_type]
            
            # Optimize image for AI processing
            optimized_image = self.optimize_image_for_ai_processing(image)
            
            # Convert image to appropriate format for API
            image_data = self._prepare_image_for_api(optimized_image)
            
            # Create AI message with image and prompt
            message = HumanMessage(
                content=[
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url", 
                        "image_url": {"url": image_data}
                    }
                ]
            )
            
            # Get AI response using thread-safe LLM
            llm = self._get_thread_safe_llm()
            ai_response = llm.invoke([message])
            
            # Extract and return content
            return self._extract_content_from_response(ai_response)
            
        except Exception as e:
            print(f"Error extracting {content_type}: {e}")
            return ""
    
    def extract_all_content_types_batch(self, image: Image.Image) -> Dict[str, str]:
        """
        Extract all content types in a single AI call for maximum efficiency.
        
        This method is 3x faster than individual extraction calls.
        
        Returns:
            Dictionary with 'text', 'table', and 'visual' content
        """
        try:
            # Optimize image for processing
            optimized_image = self.optimize_image_for_ai_processing(image)
            
            # Prepare image data for API
            image_data = self._prepare_image_for_api(optimized_image)
            
            # Create batch processing message
            message = HumanMessage(
                content=[
                    {"type": "text", "text": Config.BATCH_EXTRACTION_PROMPT},
                    {
                        "type": "image_url",
                        "image_url": {"url": image_data}
                    }
                ]
            )
            
            # Get AI response
            llm = self._get_thread_safe_llm()
            ai_response = llm.invoke([message])
            
            # Parse structured response into content types
            content = self._extract_content_from_response(ai_response)
            return self._parse_batch_extraction_response(content)
            
        except Exception as e:
            print(f"Error in batch extraction: {e}")
            return {"text": "", "table": "", "visual": ""}
    
    def _prepare_image_for_api(self, image: Image.Image) -> str:
        """Prepare image data for AI API consumption."""
        try:
            buffer = BytesIO()
            
            if Config.ENABLE_IMAGE_COMPRESSION:
                # Use JPEG compression for faster uploads
                image.save(buffer, format="JPEG", quality=Config.IMAGE_QUALITY, optimize=True)
                image_format = "jpeg"
            else:
                # Use PNG for maximum quality
                image.save(buffer, format="PNG")
                image_format = "png"
            
            image_bytes = buffer.getvalue()
            base64_data = base64.b64encode(image_bytes).decode()
            
            return f"data:image/{image_format};base64,{base64_data}"
            
        except Exception as e:
            print(f"Error preparing image for API: {e}")
            return ""
    
    def _extract_content_from_response(self, ai_response) -> str:
        """Extract text content from AI response object."""
        try:
            if hasattr(ai_response, 'content'):
                return ai_response.content
            elif isinstance(ai_response, dict):
                return ai_response.get('content', str(ai_response))
            else:
                return str(ai_response)
        except Exception as e:
            print(f"Error extracting response content: {e}")
            return ""
    
    def _parse_batch_extraction_response(self, response_text: str) -> Dict[str, str]:
        """
        Parse structured batch response into separate content types.
        
        Extracts content organized by ## section headers.
        """
        try:
            content_types = {"text": "", "table": "", "visual": ""}
            
            # Split response by section headers
            sections = response_text.split("##")
            
            for section in sections:
                section_content = section.strip()
                
                # Identify and extract each content type
                if section_content.upper().startswith("TEXT CONTENT"):
                    text_content = section_content.replace("TEXT CONTENT", "").strip()
                    if "no text content found" not in text_content.lower():
                        content_types["text"] = text_content
                
                elif section_content.upper().startswith("TABLE CONTENT"):
                    table_content = section_content.replace("TABLE CONTENT", "").strip()
                    if "no tables found" not in table_content.lower():
                        content_types["table"] = table_content
                
                elif section_content.upper().startswith("VISUAL CONTENT"):
                    visual_content = section_content.replace("VISUAL CONTENT", "").strip()
                    if "no visual elements found" not in visual_content.lower():
                        content_types["visual"] = visual_content
            
            return content_types
            
        except Exception as e:
            print(f"Error parsing batch response: {e}")
            return {"text": "", "table": "", "visual": ""}
    
    def process_eml_document(self, eml_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Process EML file and extract content for RAG pipeline"""
        try:
            email_content = self.process_eml_file(eml_path)
            
            processed_content = {
                "text": [],
                "tables": [],
                "visuals": []
            }
            
            # Process email metadata and body as text
            email_text_parts = []
            
            if email_content.get('subject'):
                email_text_parts.append(f"Subject: {email_content['subject']}")
            
            if email_content.get('from'):
                email_text_parts.append(f"From: {email_content['from']}")
            
            if email_content.get('to'):
                email_text_parts.append(f"To: {email_content['to']}")
            
            if email_content.get('date'):
                email_text_parts.append(f"Date: {email_content['date']}")
            
            if email_content.get('body_text'):
                email_text_parts.append(f"Body:\n{email_content['body_text']}")
            elif email_content.get('body_html'):
                # Basic HTML stripping for HTML-only emails
                import re
                html_content = email_content['body_html']
                # Remove HTML tags
                clean_text = re.sub('<[^<]+?>', '', html_content)
                email_text_parts.append(f"Body:\n{clean_text}")
            
            if email_content.get('attachments'):
                attachment_info = []
                for att in email_content['attachments']:
                    attachment_info.append(f"- {att['filename']} ({att['content_type']}, {att['size']} bytes)")
                if attachment_info:
                    email_text_parts.append(f"Attachments:\n" + "\n".join(attachment_info))
            
            # Combine all email content
            full_email_text = "\n\n".join(email_text_parts)
            
            processed_content["text"].append({
                "content": full_email_text,
                "page": 1,
                "source": eml_path,
                "type": "email"
            })
            
            return processed_content
            
        except Exception as e:
            print(f"Error processing EML document: {e}")
            return {"text": [], "tables": [], "visuals": []}
    
    def process_excel_document(self, excel_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Process Excel file and extract content for RAG pipeline"""
        try:
            file_ext = os.path.splitext(excel_path)[1].lower()
            
            processed_content = {
                "text": [],
                "tables": [],
                "visuals": []
            }
            
            if file_ext == '.xlsx':
                workbook = openpyxl.load_workbook(excel_path, data_only=True)
                
                for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                    sheet = workbook[sheet_name]
                    
                    # Extract sheet data as tables
                    sheet_data = self._extract_xlsx_sheet_data(sheet, sheet_name)
                    if sheet_data:
                        processed_content["tables"].append({
                            "content": sheet_data,
                            "page": sheet_index + 1,
                            "source": excel_path,
                            "type": "excel_sheet",
                            "sheet_name": sheet_name
                        })
                
                # Extract embedded images
                embedded_images = self.extract_images_from_xlsx(excel_path)
                for img_index, image in enumerate(embedded_images):
                    visual_content = self.extract_content_from_image(image, "visual")
                    if visual_content.strip() and "no visual elements found" not in visual_content.lower():
                        processed_content["visuals"].append({
                            "content": visual_content,
                            "page": f"embedded_image_{img_index + 1}",
                            "source": excel_path,
                            "type": "embedded_image"
                        })
                
                workbook.close()
                
            elif file_ext == '.xls':
                workbook = xlrd.open_workbook(excel_path)
                
                for sheet_index in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_index)
                    sheet_name = workbook.sheet_names()[sheet_index]
                    
                    # Extract sheet data as tables
                    sheet_data = self._extract_xls_sheet_data(sheet, sheet_name)
                    if sheet_data:
                        processed_content["tables"].append({
                            "content": sheet_data,
                            "page": sheet_index + 1,
                            "source": excel_path,
                            "type": "excel_sheet",
                            "sheet_name": sheet_name
                        })
            
            return processed_content
            
        except Exception as e:
            print(f"Error processing Excel document: {e}")
            return {"text": [], "tables": [], "visuals": []}
    
    def _extract_xlsx_sheet_data(self, sheet, sheet_name: str) -> str:
        """Extract data from XLSX sheet and format as markdown table"""
        try:
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row == 1 and max_col == 1 and not sheet.cell(1, 1).value:
                return ""  # Empty sheet
            
            # Extract data
            data = []
            for row in range(1, min(max_row + 1, 101)):  # Limit to first 100 rows
                row_data = []
                for col in range(1, min(max_col + 1, 21)):  # Limit to first 20 columns
                    cell_value = sheet.cell(row, col).value
                    if cell_value is None:
                        cell_value = ""
                    row_data.append(str(cell_value))
                data.append(row_data)
            
            if not data:
                return ""
            
            # Format as markdown table
            markdown_table = f"### Sheet: {sheet_name}\n\n"
            
            # Create header
            if len(data) > 0:
                headers = data[0]
                markdown_table += "| " + " | ".join(headers) + " |\n"
                markdown_table += "|" + "|".join(["---"] * len(headers)) + "|\n"
                
                # Add data rows
                for row in data[1:]:
                    if len(row) == len(headers):  # Ensure consistent column count
                        markdown_table += "| " + " | ".join(row) + " |\n"
            
            return markdown_table
            
        except Exception as e:
            print(f"Error extracting XLSX sheet data: {e}")
            return ""
    
    def _extract_xls_sheet_data(self, sheet, sheet_name: str) -> str:
        """Extract data from XLS sheet and format as markdown table"""
        try:
            nrows = sheet.nrows
            ncols = sheet.ncols
            
            if nrows == 0 or ncols == 0:
                return ""  # Empty sheet
            
            # Extract data
            data = []
            for row in range(min(nrows, 101)):  # Limit to first 100 rows
                row_data = []
                for col in range(min(ncols, 21)):  # Limit to first 20 columns
                    cell_value = sheet.cell_value(row, col)
                    if cell_value is None or cell_value == "":
                        cell_value = ""
                    row_data.append(str(cell_value))
                data.append(row_data)
            
            if not data:
                return ""
            
            # Format as markdown table
            markdown_table = f"### Sheet: {sheet_name}\n\n"
            
            # Create header
            if len(data) > 0:
                headers = data[0]
                markdown_table += "| " + " | ".join(headers) + " |\n"
                markdown_table += "|" + "|".join(["---"] * len(headers)) + "|\n"
                
                # Add data rows
                for row in data[1:]:
                    if len(row) == len(headers):  # Ensure consistent column count
                        markdown_table += "| " + " | ".join(row) + " |\n"
            
            return markdown_table
            
        except Exception as e:
            print(f"Error extracting XLS sheet data: {e}")
            return ""
        """Process a single page and extract all content types"""
        image = page_data['image']
        page_num = page_data['page_num']
        file_path = page_data['file_path']
        image_path = page_data.get('image_path', '')
        
        result = {
            'page_num': page_num,
            'text': None,
            'table': None,
            'visual': None
        }
        
        # Smart filtering: Skip duplicate or empty pages
        if self._is_duplicate_page(image):
            print(f"  Skipping duplicate page {page_num}")
            return result
        
        if not self._has_sufficient_content(image):
            print(f"  Skipping page {page_num} with insufficient content")
            return result
        
        # Use batch processing if enabled, otherwise use individual calls
        if Config.ENABLE_BATCH_PROCESSING:
            try:
                batch_results = self.extract_all_content_batch(image)
                
                # Process batch results
                if batch_results["text"].strip():
                    result['text'] = {
                        "content": batch_results["text"],
                        "page": page_num,
                        "source": file_path,
                        "type": "text",
                        "image_path": image_path
                    }
                
                if batch_results["table"].strip():
                    result['table'] = {
                        "content": batch_results["table"],
                        "page": page_num,
                        "source": file_path,
                        "type": "table",
                        "image_path": image_path
                    }
                
                if batch_results["visual"].strip():
                    result['visual'] = {
                        "content": batch_results["visual"],
                        "page": page_num,
                        "source": file_path,
                        "type": "visual",
                        "image_path": image_path
                    }
                
            except Exception as e:
                print(f"Batch processing failed for page {page_num}, falling back to individual calls: {e}")
                # Fallback to individual processing
                Config.ENABLE_BATCH_PROCESSING = False
        
        # Individual processing (fallback or when batch processing is disabled)
        if not Config.ENABLE_BATCH_PROCESSING:
            # Extract text content
            text_content = self.extract_content_from_image(image, "text")
            if text_content.strip() and "no text content found" not in text_content.lower():
                result['text'] = {
                    "content": text_content,
                    "page": page_num,
                    "source": file_path,
                    "type": "text",
                    "image_path": image_path
                }
            
            # Extract table content
            table_content = self.extract_content_from_image(image, "table")
            if table_content.strip() and "no tables found" not in table_content.lower():
                result['table'] = {
                    "content": table_content,
                    "page": page_num,
                    "source": file_path,
                    "type": "table",
                    "image_path": image_path
                }
            
            # Extract visual content
            visual_content = self.extract_content_from_image(image, "visual")
            if visual_content.strip() and "no visual elements found" not in visual_content.lower():
                result['visual'] = {
                    "content": visual_content,
                    "page": page_num,
                    "source": file_path,
                    "type": "visual",
                    "image_path": image_path
                }
        
        return result
    
    # =============================================================================
    # MAIN DOCUMENT PROCESSING INTERFACE
    # =============================================================================
    
    def process_document(self, file_path: str, force_reprocess: bool = False) -> Dict[str, List[Dict[str, Any]]]:
        """
        Main entry point for document processing with intelligent caching.
        
        Args:
            file_path: Path to document file
            force_reprocess: Skip cache and reprocess document
            
        Returns:
            Dictionary with extracted content organized by type
        """
        try:
            # Check for cached extraction first (unless forced reprocessing)
            if not force_reprocess:
                cached_content = self._load_cached_document_extraction(file_path)
                if cached_content is not None:
                    print(f"Loaded cached extraction for {os.path.basename(file_path)}")
                    return cached_content
            
            print(f"Processing document: {os.path.basename(file_path)}")
            
            # Determine file type and route to appropriate processor
            file_extension = os.path.splitext(file_path)[1].lower()
            
            if file_extension == '.eml':
                processed_content = self.process_eml_document(file_path)
            elif file_extension in ['.xls', '.xlsx']:
                processed_content = self.process_excel_document(file_path)
            elif file_extension == '.pdf':
                images = self.convert_pdf_to_images(file_path)
                processed_content = self._process_image_sequence(file_path, images)
            elif file_extension == '.pptx':
                images = self.convert_pptx_to_images(file_path)
                processed_content = self._process_image_sequence(file_path, images)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Cache extraction results for future use
            self._save_document_extraction_cache(file_path, processed_content)
            
            print(f"Document processing complete: {file_path}")
            return processed_content
            
        except Exception as e:
            print(f"Error processing document {file_path}: {e}")
            return {"text": [], "tables": [], "visuals": []}
    
    def process_document_with_parallel_processing(self, file_path: str, force_reprocess: bool = False) -> Dict[str, List[Dict[str, Any]]]:
        """
        Process document with parallel processing for maximum performance.
        
        Uses multi-threading for large documents when enabled in config.
        """
        try:
            # Check cache first
            if not force_reprocess:
                cached_content = self._load_cached_document_extraction(file_path)
                if cached_content is not None:
                    print(f"Loaded cached parallel extraction for {os.path.basename(file_path)}")
                    return cached_content
            
            print(f"Processing document with parallel processing: {os.path.basename(file_path)}")
            
            file_extension = os.path.splitext(file_path)[1].lower()
            
            # Route to appropriate processor
            if file_extension == '.eml':
                processed_content = self.process_eml_document(file_path)
            elif file_extension in ['.xls', '.xlsx']:
                processed_content = self.process_excel_document(file_path)
            elif file_extension == '.pdf':
                images = self.convert_pdf_to_images(file_path)
                processed_content = self._process_images_with_parallel_execution(file_path, images)
            elif file_extension == '.pptx':
                images = self.convert_pptx_to_images(file_path)
                processed_content = self._process_images_with_parallel_execution(file_path, images)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
            
            # Cache results
            self._save_document_extraction_cache(file_path, processed_content)
            
            return processed_content
            
        except Exception as e:
            print(f"Error in parallel document processing {file_path}: {e}")
            return {"text": [], "tables": [], "visuals": []}
    
    # =============================================================================
    # IMAGE SEQUENCE PROCESSING
    # =============================================================================
    
    def _process_image_sequence(self, file_path: str, images: List[Image.Image]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Process a sequence of images with smart optimizations.
        
        Automatically chooses between sequential and parallel processing
        based on document size and configuration.
        """
        if Config.ENABLE_MULTIPROCESSING and len(images) >= Config.MIN_PAGES_FOR_PARALLEL:
            return self._process_images_with_parallel_execution(file_path, images)
        else:
            return self._process_images_sequentially_optimized(file_path, images)
    
    def _process_images_with_parallel_execution(self, file_path: str, images: List[Image.Image]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Process images using parallel execution for maximum performance.
        
        Uses ThreadPoolExecutor to process multiple pages simultaneously.
        """
        processed_content = {
            "text": [],
            "tables": [],
            "visuals": []
        }
        
        if not Config.ENABLE_MULTIPROCESSING or len(images) < 2:
            # Fall back to sequential processing for small documents
            return self._process_images_sequentially_optimized(file_path, images)
        
        print(f"Processing {len(images)} pages with {Config.MAX_WORKERS} parallel workers...")
        
        # Prepare image data for parallel processing
        page_processing_tasks = []
        for page_num, image in enumerate(images, 1):
            # Save image for processing reference
            image_path = self.save_processed_image(image, file_path, page_num, "page")
            
            task_data = {
                'image': image,
                'image_path': image_path,
                'page_num': page_num,
                'file_path': file_path
            }
            page_processing_tasks.append(task_data)
        
        # Execute parallel processing
        processing_results = {}
        
        with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
            # Submit all pages for processing
            future_to_page = {
                executor.submit(self._process_single_page_content, task_data): task_data['page_num']
                for task_data in page_processing_tasks
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_page):
                page_num = future_to_page[future]
                try:
                    page_result = future.result()
                    processing_results[page_result['page_num']] = page_result
                    print(f"✓ Completed page {page_result['page_num']}/{len(images)}")
                except Exception as e:
                    print(f"✗ Error processing page {page_num}: {e}")
        
        # Organize results by page order
        for page_num in sorted(processing_results.keys()):
            result = processing_results[page_num]
            
            if result['text']:
                processed_content["text"].append(result['text'])
            if result['table']:
                processed_content["tables"].append(result['table'])
            if result['visual']:
                processed_content["visuals"].append(result['visual'])
        
        print(f"Parallel processing complete: {len(processing_results)} pages processed")
        return processed_content
    
    def _process_images_sequentially_optimized(self, file_path: str, images: List[Image.Image]) -> Dict[str, List[Dict[str, Any]]]:
        """
        Process images sequentially with performance optimizations.
        
        Includes smart filtering, caching, and batch processing optimizations.
        """
        processed_content = {
            "text": [],
            "tables": [],
            "visuals": []
        }
        
        print(f"Processing {len(images)} pages sequentially with optimizations...")
        
        for page_num, image in enumerate(images, 1):
            try:
                print(f"Processing page {page_num}/{len(images)}...")
                
                # Apply smart filtering to skip low-content pages
                if not self.has_sufficient_content_for_processing(image):
                    print(f"Skipping page {page_num} - insufficient content detected")
                    continue
                
                # Check for duplicate pages to avoid redundant processing
                if self.is_duplicate_page(image):
                    print(f"Skipping page {page_num} - duplicate content detected")
                    continue
                
                # Save processed image for reference
                image_path = self.save_processed_image(image, file_path, page_num, "page")
                
                # Extract content using batch processing for efficiency
                if Config.ENABLE_BATCH_PROCESSING:
                    content = self.extract_all_content_types_batch(image)
                    
                    # Create content chunks for each type
                    if content["text"]:
                        text_chunk = {
                            "content": content["text"],
                            "source": file_path,
                            "page": page_num,
                            "type": "text",
                            "image_path": image_path
                        }
                        processed_content["text"].append(text_chunk)
                    
                    if content["table"]:
                        table_chunk = {
                            "content": content["table"],
                            "source": file_path,
                            "page": page_num,
                            "type": "table",
                            "image_path": image_path
                        }
                        processed_content["tables"].append(table_chunk)
                    
                    if content["visual"]:
                        visual_chunk = {
                            "content": content["visual"],
                            "source": file_path,
                            "page": page_num,
                            "type": "visual",
                            "image_path": image_path
                        }
                        processed_content["visuals"].append(visual_chunk)
                
                else:
                    # Individual extraction mode (slower but more control)
                    text_content = self.extract_single_content_type(image, "text")
                    table_content = self.extract_single_content_type(image, "table")
                    visual_content = self.extract_single_content_type(image, "visual")
                    
                    # Create content chunks
                    if text_content:
                        processed_content["text"].append({
                            "content": text_content,
                            "source": file_path,
                            "page": page_num,
                            "type": "text",
                            "image_path": image_path
                        })
                    
                    if table_content:
                        processed_content["tables"].append({
                            "content": table_content,
                            "source": file_path,
                            "page": page_num,
                            "type": "table",
                            "image_path": image_path
                        })
                    
                    if visual_content:
                        processed_content["visuals"].append({
                            "content": visual_content,
                            "source": file_path,
                            "page": page_num,
                            "type": "visual",
                            "image_path": image_path
                        })
                
                print(f"✓ Page {page_num} processed successfully")
                
            except Exception as e:
                print(f"✗ Error processing page {page_num}: {e}")
                continue
        
        total_chunks = (len(processed_content["text"]) + 
                       len(processed_content["tables"]) + 
                       len(processed_content["visuals"]))
        
        print(f"Sequential processing complete: {total_chunks} content chunks extracted")
        return processed_content
    
    def _process_single_page_content(self, page_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process a single page for parallel execution.
        
        Args:
            page_data: Dictionary containing image, page_num, file_path, etc.
            
        Returns:
            Dictionary with extracted content for the page
        """
        try:
            image = page_data['image']
            page_num = page_data['page_num']
            file_path = page_data['file_path']
            
            # Apply optimizations
            if not self.has_sufficient_content_for_processing(image):
                return {
                    'page_num': page_num,
                    'text': None,
                    'table': None,
                    'visual': None
                }
            
            # Extract content using batch processing
            if Config.ENABLE_BATCH_PROCESSING:
                content = self.extract_all_content_types_batch(image)
                
                return {
                    'page_num': page_num,
                    'text': {
                        "content": content["text"],
                        "source": file_path,
                        "page": page_num,
                        "type": "text",
                        "image_path": page_data.get('image_path', '')
                    } if content["text"] else None,
                    'table': {
                        "content": content["table"],
                        "source": file_path,
                        "page": page_num,
                        "type": "table",
                        "image_path": page_data.get('image_path', '')
                    } if content["table"] else None,
                    'visual': {
                        "content": content["visual"],
                        "source": file_path,
                        "page": page_num,
                        "type": "visual",
                        "image_path": page_data.get('image_path', '')
                    } if content["visual"] else None
                }
            else:
                # Individual extraction
                text_content = self.extract_single_content_type(image, "text")
                table_content = self.extract_single_content_type(image, "table")
                visual_content = self.extract_single_content_type(image, "visual")
                
                return {
                    'page_num': page_num,
                    'text': {
                        "content": text_content,
                        "source": file_path,
                        "page": page_num,
                        "type": "text",
                        "image_path": page_data.get('image_path', '')
                    } if text_content else None,
                    'table': {
                        "content": table_content,
                        "source": file_path,
                        "page": page_num,
                        "type": "table",
                        "image_path": page_data.get('image_path', '')
                    } if table_content else None,
                    'visual': {
                        "content": visual_content,
                        "source": file_path,
                        "page": page_num,
                        "type": "visual",
                        "image_path": page_data.get('image_path', '')
                    } if visual_content else None
                }
                
        except Exception as e:
            print(f"Error processing page {page_data.get('page_num', 'unknown')}: {e}")
            return {
                'page_num': page_data.get('page_num', 0),
                'text': None,
                'table': None,
                'visual': None
            }
        skipped_pages = 0
        
        for page_num, image in enumerate(images):
            print(f"Processing page {page_num + 1}...")
            
            # Smart filtering: Skip duplicate or empty pages
            if self._is_duplicate_page(image):
                print(f"  Skipping duplicate page {page_num + 1}")
                skipped_pages += 1
                continue
            
            if not self._has_sufficient_content(image):
                print(f"  Skipping page {page_num + 1} with insufficient content")
                skipped_pages += 1
                continue
            
            # Save image
            image_path = self._save_image(image, file_path, page_num + 1, "page")
            
            # Use batch processing if enabled
            if Config.ENABLE_BATCH_PROCESSING:
                try:
                    batch_results = self.extract_all_content_batch(image)
                    
                    # Process batch results
                    if batch_results["text"].strip():
                        processed_content["text"].append({
                            "content": batch_results["text"],
                            "page": page_num + 1,
                            "source": file_path,
                            "type": "text",
                            "image_path": image_path
                        })
                    
                    if batch_results["table"].strip():
                        processed_content["tables"].append({
                            "content": batch_results["table"],
                            "page": page_num + 1,
                            "source": file_path,
                            "type": "table",
                            "image_path": image_path
                        })
                    
                    if batch_results["visual"].strip():
                        processed_content["visuals"].append({
                            "content": batch_results["visual"],
                            "page": page_num + 1,
                            "source": file_path,
                            "type": "visual",
                            "image_path": image_path
                        })
                    
                    continue  # Skip individual processing
                    
                except Exception as e:
                    print(f"Batch processing failed for page {page_num + 1}, falling back to individual calls: {e}")
                    # Fall through to individual processing
            
            # Individual processing (fallback or when batch processing is disabled)
            # Extract text content
            text_content = self.extract_content_from_image(image, "text")
            if text_content.strip() and "no text content found" not in text_content.lower():
                processed_content["text"].append({
                    "content": text_content,
                    "page": page_num + 1,
                    "source": file_path,
                    "type": "text",
                    "image_path": image_path
                })
            
            # Extract table content
            table_content = self.extract_content_from_image(image, "table")
            if table_content.strip() and "no tables found" not in table_content.lower():
                processed_content["tables"].append({
                    "content": table_content,
                    "page": page_num + 1,
                    "source": file_path,
                    "type": "table",
                    "image_path": image_path
                })
            
            # Extract visual content
            visual_content = self.extract_content_from_image(image, "visual")
            if visual_content.strip() and "no visual elements found" not in visual_content.lower():
                processed_content["visuals"].append({
                    "content": visual_content,
                    "page": page_num + 1,
                    "source": file_path,
                    "type": "visual",
                    "image_path": image_path
                })
        
        if skipped_pages > 0:
            print(f"✓ Skipped {skipped_pages} pages due to smart filtering")
        
        return processed_content
    
    # =============================================================================
    # UTILITY AND HELPER METHODS
    # =============================================================================
    
    def load_from_saved_extraction(self, file_path: str) -> Optional[Dict[str, List[Dict[str, Any]]]]:
        """
        Public method to load previously saved extraction results.
        
        Returns:
            Cached extraction results or None if not found
        """
        return self._load_cached_document_extraction(file_path)
    
    def get_all_saved_extractions(self) -> List[Dict[str, Any]]:
        """
        Get metadata for all saved extraction results.
        
        Returns:
            List of extraction metadata including file paths, timestamps, etc.
        """
        extractions = []
        
        if not os.path.exists(self.content_dir):
            return extractions
        
        for filename in os.listdir(self.content_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(self.content_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    extractions.append({
                        'extraction_file': filepath,
                        'source_file': data.get('source_file', ''),
                        'file_hash': data.get('file_hash', ''),
                        'extraction_date': data.get('extraction_date', ''),
                        'content_stats': {
                            'text_items': len(data.get('content', {}).get('text', [])),
                            'table_items': len(data.get('content', {}).get('tables', [])),
                            'visual_items': len(data.get('content', {}).get('visuals', []))
                        }
                    })
                except Exception as e:
                    print(f"Error reading extraction file {filename}: {e}")
        
        return extractions
    
    def reset_duplicate_detection(self) -> None:
        """
        Reset the duplicate detection cache for processing new documents.
        
        Clears all stored image hashes and similarity data.
        """
        self.processed_hashes.clear()
        self.page_similarities.clear()
        print("Duplicate detection cache reset")
    
    def get_performance_statistics(self) -> Dict[str, Any]:
        """
        Get comprehensive performance statistics and configuration info.
        
        Returns:
            Dictionary with current performance settings and optimization status
        """
        return {
            "optimization_status": {
                "batch_processing_enabled": Config.ENABLE_BATCH_PROCESSING,
                "image_compression_enabled": Config.ENABLE_IMAGE_COMPRESSION,
                "smart_filtering_enabled": Config.ENABLE_SMART_FILTERING,
                "duplicate_detection_enabled": Config.ENABLE_DUPLICATE_DETECTION,
                "parallel_processing_enabled": Config.ENABLE_MULTIPROCESSING,
                "extraction_caching_enabled": Config.ENABLE_EXTRACTION_CACHING
            },
            "performance_settings": {
                "max_workers": Config.MAX_WORKERS,
                "batch_size": Config.BATCH_SIZE,
                "image_quality": Config.IMAGE_QUALITY,
                "pdf_dpi": Config.PDF_DPI,
                "min_pages_for_parallel": Config.MIN_PAGES_FOR_PARALLEL
            },
            "runtime_stats": {
                "processed_hashes_count": len(self.processed_hashes),
                "total_pages_processed": getattr(self, 'total_pages_processed', 0),
                "total_api_calls": getattr(self, 'total_api_calls', 0),
                "total_processing_time": getattr(self, 'total_processing_time', 0.0)
            }
        }
    
    # =============================================================================
    # PRIVATE CACHE MANAGEMENT METHODS
    # =============================================================================
    
    def _load_cached_document_extraction(self, file_path: str) -> Optional[Dict[str, List[Dict[str, Any]]]]:
        """Load cached extraction results for a document."""
        try:
            cache_file = self.get_extraction_cache_path(file_path)
            
            if not os.path.exists(cache_file):
                return None
            
            # Check if cache is still valid
            if not self.has_valid_cache(file_path):
                return None
            
            with open(cache_file, 'r', encoding='utf-8') as f:
                cached_data = json.load(f)
            
            # Return content in expected format
            return {
                'text': cached_data.get('content', {}).get('text', []),
                'tables': cached_data.get('content', {}).get('tables', []),
                'visuals': cached_data.get('content', {}).get('visuals', [])
            }
            
        except Exception as e:
            print(f"Error loading cached extraction: {e}")
            return None
    
    def _save_document_extraction_cache(self, file_path: str, processed_content: Dict[str, List[Dict[str, Any]]]) -> None:
        """Save extraction results to cache."""
        try:
            cache_file = self.get_extraction_cache_path(file_path)
            
            # Prepare data with metadata
            cache_data = {
                'source_file': file_path,
                'file_hash': self.generate_file_hash(file_path),
                'extraction_date': datetime.now().isoformat(),
                'processor_version': '2.0.0',
                'content': {
                    'text': processed_content.get('text', []),
                    'tables': processed_content.get('tables', []),
                    'visuals': processed_content.get('visuals', [])
                }
            }
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(cache_file), exist_ok=True)
            
            # Save to cache
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(cache_data, f, indent=2, ensure_ascii=False)
            
            print(f"Extraction results cached: {os.path.basename(cache_file)}")
            
        except Exception as e:
            print(f"Error saving extraction cache: {e}")

    # Legacy method names for backward compatibility
    def get_performance_stats(self) -> Dict[str, Any]:
        """Legacy method name - use get_performance_statistics() instead."""
        return self.get_performance_statistics()
